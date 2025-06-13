import streamlit as st
from datetime import datetime, timedelta
from io import BytesIO
from fpdf import FPDF
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from collections import defaultdict

st.set_page_config(page_title="Schedule Planner", layout="centered")
st.title("🗓️ Schedule Planner Web App")

# === Helper Functions ===

def parse_time(time_str):
    time_str = time_str.strip().lower().replace(" ", "")
    try:
        return datetime.strptime(time_str, "%I%p")         # e.g., "12pm"
    except:
        pass
    try:
        return datetime.strptime(time_str, "%I:%M%p")      # e.g., "12:00pm"
    except:
        pass
    try:
        return datetime.strptime(time_str, "%H:%M")        # e.g., "12:00" (24hr)
    except:
        pass
    return None

def get_minutes(start, end):
    delta = end - start
    return max(0, int(delta.total_seconds() // 60))
    
def get_week_number(start_date, current_date):
    start_sunday = start_date - timedelta(days=start_date.weekday() + 1 if start_date.weekday() != 6 else 0)
    current_sunday = current_date - timedelta(days=current_date.weekday() + 1 if current_date.weekday() != 6 else 0)
    return ((current_sunday - start_sunday).days // 7) + 1
    
def generate_pdf(data):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Work Schedule Report", ln=True, align='C')
    pdf.set_font("Arial", "", 12)
    pdf.ln(5)

    # Group data by week
    grouped_weeks = defaultdict(list)
    week_totals = defaultdict(int)
    grand_total = 0

    # for row in schedule_data:
    for row in data:
        week = row['week']
        grouped_weeks[week].append(row)
        week_totals[week] += row['duration']
        grand_total += row['duration']

    for week_num in sorted(grouped_weeks):
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, f"Week {week_num}", ln=True)
        pdf.set_font("Arial", "B", 12)
        pdf.cell(30, 10, "Day", 1)
        pdf.cell(30, 10, "Date", 1)
        pdf.cell(40, 10, "Time In", 1)
        pdf.cell(40, 10, "Time Out", 1)
        pdf.cell(50, 10, "Duration", 1)
        pdf.ln()

        pdf.set_font("Arial", "", 11)
        for row in grouped_weeks[week_num]:
            pdf.cell(30, 10, row['day'], 1)
            pdf.cell(30, 10, row['date'], 1)
            pdf.cell(40, 10, row['time_in'], 1)
            pdf.cell(40, 10, row['time_out'], 1)
            duration = f"{row['duration'] // 60} hrs {row['duration'] % 60} min"
            pdf.cell(50, 10, duration, 1)
            pdf.ln()

        week_total = week_totals[week_num]
        total_str = f"Week {week_num} Total: {week_total // 60} hrs {week_total % 60} min"
        pdf.cell(0, 10, total_str, ln=True)
        pdf.ln(4)

    # Final Total
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"FINAL TOTAL: {grand_total // 60} hrs {grand_total % 60} min", ln=True, align='C')
    
    # Proper in-memory PDF output
    pdf_bytes = pdf.output(dest='S').encode('latin1') # <- Important fix
    return BytesIO(pdf_bytes)

def generate_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(["Date", "Time In", "Time Out", "Minutes Worked"])

    for row in data:
        ws.append([row["date"], row["time_in"], row["time_out"], row["duration"]])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_csv(data):
    buffer = BytesIO()
    buffer.write("Date,Time In,Time Out,Minutes Worked\n".encode('utf-8'))
    for row in data:
        line = f"{row['date']},{row['time_in']},{row['time_out']},{row['duration']}\n"
        buffer.write(line.encode('utf-8'))
    buffer.seek(0)
    return buffer

def generate_json(data):
    buffer = BytesIO()
    buffer.write(json.dumps(data, indent=2).encode('utf-8'))
    buffer.seek(0)
    return buffer

st.subheader("Enter Your Work Schedule")
st.caption("⏱️ Time Format: Use 12-hour (e.g., 9:00AM or 4:30PM) or 24-hour (e.g., 13:00)")

# === UI Inputs ===

schedule_data = []

date_range = st.date_input("Select Date Range", [datetime.today(), datetime.today() + timedelta(days=5)])

# ✅ Define checkboxes before using them
use_typical = st.checkbox("Use Typical Hours for Weekdays?")
include_weekends = st.checkbox("Include weekends in schedule?")

col1, col2 = st.columns(2)

if use_typical:
    with col1:
        default_in = st.text_input("Typical Time In", "9:00AM")
    with col2:
        default_out = st.text_input("Typical Time Out", "5:00PM")

if date_range and len(date_range) == 2:
    start_date, end_date = date_range
    
    current = start_date

    while current <= end_date:
        day_name = current.strftime('%A')

        if not include_weekends and day_name in ["Saturday", "Sunday"]:
            current += timedelta(days=1)
            continue

        st.markdown(f"**{day_name}, {current.strftime('%m/%d/%Y')}**")

        entry_index = 1
        add_another = True
        
        #add_extra = st.checkbox(f"➕ Add another entry for {current.strftime('%m/%d/%Y')}?", key=f"extra_checkbox_{current}")

        while add_another:
                # time_in = st.text_input(f"Time In ({entry_index}) - {current.strftime('%m/%d/%Y')}", key=f"in_{current}_{entry_index}")
                # time_out = st.text_input(f"Time Out ({entry_index}) - {current.strftime('%m/%d/%Y')}", key=f"out_{current}_{entry_index}")
                col1, col2 = st.columns(2)
            
                if use_typical and current.weekday() < 5 and entry_index == 1:
                    time_in = default_in
                    time_out = default_out
                    # st.markdown(f"Auto-filled: {time_in} to {time_out}")
                else:
                    with col1:
                        st.text_input("Auto-filled Time In", value=time_in, key=f"in_{current}_{entry_index}", disabled=True)
                    with col2:
                        st.text_input("Auto-filled Time Out", value=time_out, key=f"out_{current}_{entry_index}", disabled=True)

                # Save auto-filled values only once
                if default_in and default_out:
                    t_in = parse_time(default_in)
                    t_out = parse_time(default_out)
                    if t_in and t_out and t_out > t_in:
                        duration = get_minutes(t_in, t_out)
                        week_number = get_week_number(start_date, current)
                        schedule_data.append({
                            "week": week_number,
                            "day": current.strftime("%A"),
                            "date": current.strftime("%m/%d/%Y"),
                            "time_in": t_in.strftime("%I:%M %p"),
                            "time_out": t_out.strftime("%I:%M %p"),
                            "duration": duration
                        })
                        add_another = st.checkbox(f"➕ Add another entry for {current.strftime('%m/%d/%Y')}?", key=f"another_{current}_{entry_index}")
                    else:
                        st.warning(f"Invalid time entry {entry_index} on {current.strftime('%m/%d/%Y')}.")
                        add_another = False
                else:
                    add_another = False
    
                entry_index += 1
                # else:
                    
                #     with col1:
                #         time_in = st.text_input(f"Time In ({entry_index}) - {current.strftime('%m/%d/%Y')}", key=f"in_{current}_{entry_index}")
                #     with col2:
                #         time_out = st.text_input(f"Time Out ({entry_index}) - {current.strftime('%m/%d/%Y')}", key=f"out_{current}_{entry_index}")

                # Only record entries with valid times
        #         if time_in and time_out:
        #             t_in = parse_time(time_in)
        #             t_out = parse_time(time_out)
        #             if t_in and t_out and t_out > t_in:
        #                 duration = get_minutes(t_in, t_out)
        #                 week_number = get_week_number(start_date, current)
        #                 schedule_data.append({
        #                     "week": week_number,
        #                     "day": current.strftime("%A"),
        #                     "date": current.strftime("%m/%d/%Y"),
        #                     "time_in": t_in.strftime("%I:%M %p"),
        #                     "time_out": t_out.strftime("%I:%M %p"),
        #                     "duration": duration
        #                 })
                        
        #                 # Show checkbox for adding another only if current entry is valid
        #                 add_another = st.checkbox(f"➕ Add another entry for {current.strftime('%m/%d/%Y')}?", key=f"another_{current}_{entry_index}")
        #             else:
        #                 st.warning(f"Invalid time entry {entry_index} on {current.strftime('%m/%d/%Y')}.")
        #                 add_another = False
        #         else:
        #             add_another = False
                
        #         entry_index += 1
        
        # current += timedelta(days=1)
        # if use_typical and current.weekday() < 5:
        #     time_in = default_in
        #     time_out = default_out
        #     st.markdown(f"Auto-filled: {time_in} to {time_out}")
        # else:
        #     time_in = st.text_input(f"Time In ({current})", key=f"in_{current}")
        #     time_out = st.text_input(f"Time Out ({current})", key=f"out_{current}")

        # if time_in and time_out:
        #     t_in = parse_time(time_in)
        #     t_out = parse_time(time_out)
        #     if t_in and t_out and t_out > t_in:
        #         if current <= end_date:
        #             duration = get_minutes(t_in, t_out)
        #             week_number = get_week_number(start_date, current)
        #             schedule_data.append({
        #                 "week": week_number,
        #                 "day": current.strftime("%A"),
        #                 "date": current.strftime("%m/%d/%Y"),
        #                 "time_in": t_in.strftime("%I:%M %p"),
        #                 "time_out": t_out.strftime("%I:%M %p"),
        #                 "duration": duration
        #             })
        #         else:
        #             st.warning(f"Invalid times on {current}. Must be valid and Time Out after Time In.")
        #     else:
        #         st.warning("Invalid time format.")
                
        # current += timedelta(days=1)  # ✅ end of while loop block

# === Output Section ===

if schedule_data:
    st.success("✅ Schedule Data Collected!")
    
    week_groups = defaultdict(list)
    week_totals = defaultdict(int)
    grand_total_minutes = 0

    for row in schedule_data:
        week_groups[row['week']].append(row)
        week_totals[row['week']] += row['duration']
        grand_total_minutes += row['duration']

    for week in sorted(week_groups.keys()):
        st.markdown(f"### 🗓️ Week {week}")
        table_rows = []
        for r in week_groups[week]:
            table_rows.append([
                r['day'], r['date'], r['time_in'], r['time_out'], f"{r['duration'] // 60} hrs {r['duration'] % 60} min"
            ])
        st.table(table_rows)
        total = week_totals[week]
        st.info(f"Week {week} Total: {total // 60} hrs {total % 60} min")

    st.success(f"🧾 Final Total: {grand_total_minutes // 60} hrs {grand_total_minutes % 60} min 🕒")

    pdf_data = generate_pdf(schedule_data)
    excel_data = generate_excel(schedule_data)

    csv_data = generate_csv(schedule_data)
    json_data = generate_json(schedule_data)

    st.download_button("Download PDF", pdf_data, file_name="schedule.pdf")
    st.download_button("Download Excel", excel_data, file_name="schedule.xlsx")

    st.download_button("Download CSV", csv_data, file_name="schedule.csv")
    st.download_button("Download JSON", json_data, file_name="schedule.json")

st.write("App started successfully")

